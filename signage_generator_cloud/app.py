import copy
from datetime import date, datetime, time, timedelta
from io import BytesIO
from typing import Dict, List, Optional, Tuple

import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.worksheet.worksheet import Worksheet

REQUIRED_PLANS = ["Plan1", "Plan2", "Plan3"]
WEEKEND_SUFFIX = " (주말)"
HEADER_KEYWORDS = {
    "date": ["날짜", "일자"],
    "alarm": ["알람", "알림"],
    "start": ["시작예정", "시작 예정", "시작시간", "시작 시각", "시작"],
    "program": ["프로그램명", "프로그램"],
    "room": ["렉처룸", "렉쳐룸", "강의실", "룸"],
}
TIME_FORMAT = "hh:mm"
DATE_FORMAT = "m월 d일"


def is_weekend(d: date) -> bool:
    return d.weekday() >= 5


def normalize(v) -> str:
    return str(v).replace(" ", "").strip().lower() if v is not None else ""


def cell_has_keyword(value, keywords: List[str]) -> bool:
    text = normalize(value)
    return any(normalize(k) in text for k in keywords)


def find_header(ws: Worksheet, max_scan_rows: int = 40) -> Tuple[int, Dict[str, int]]:
    best_row = 1
    best_cols: Dict[str, int] = {}
    best_score = 0
    scan_rows = min(ws.max_row, max_scan_rows)
    for row_idx in range(1, scan_rows + 1):
        cols: Dict[str, int] = {}
        for cell in ws[row_idx]:
            for key, keywords in HEADER_KEYWORDS.items():
                if key not in cols and cell_has_keyword(cell.value, keywords):
                    cols[key] = cell.column
        score = len(cols)
        if score > best_score:
            best_row = row_idx
            best_cols = cols
            best_score = score
    if best_score < 3:
        raise ValueError(f"'{ws.title}' 시트에서 날짜/알람/시작/프로그램/렉처룸 헤더를 찾지 못했습니다.")
    return best_row, best_cols


def row_has_schedule_data(ws: Worksheet, row_idx: int, cols: Dict[str, int]) -> bool:
    program_col = cols.get("program")
    start_col = cols.get("start")
    room_col = cols.get("room")
    program = ws.cell(row_idx, program_col).value if program_col else None
    start = ws.cell(row_idx, start_col).value if start_col else None
    room = ws.cell(row_idx, room_col).value if room_col else None
    return bool(program and start and room)


def source_data_rows(ws: Worksheet, header_row: int, cols: Dict[str, int]) -> List[int]:
    rows: List[int] = []
    for r in range(header_row + 1, ws.max_row + 1):
        if row_has_schedule_data(ws, r, cols):
            rows.append(r)
    return rows


def last_used_row(ws: Worksheet, cols: Optional[Dict[str, int]] = None) -> int:
    min_col = min(cols.values()) if cols else 1
    max_col = max(cols.values()) if cols else ws.max_column
    for r in range(ws.max_row, 0, -1):
        for c in range(min_col, max_col + 1):
            if ws.cell(r, c).value not in (None, ""):
                return r
    return 1


def clear_schedule_body(ws: Worksheet, header_row: int, cols: Dict[str, int]) -> None:
    start_row = header_row + 1
    max_col = max(ws.max_column, max(cols.values()))
    for r in range(start_row, ws.max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(r, c).value = None


def copy_row_style_and_values(src_ws: Worksheet, dst_ws: Worksheet, src_row: int, dst_row: int, max_col: int) -> None:
    dst_ws.row_dimensions[dst_row].height = src_ws.row_dimensions[src_row].height
    for c in range(1, max_col + 1):
        src = src_ws.cell(src_row, c)
        dst = dst_ws.cell(dst_row, c)
        dst.value = src.value
        if src.has_style:
            dst._style = copy.copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.comment:
            dst.comment = copy.copy(src.comment)
        if src.hyperlink:
            dst._hyperlink = copy.copy(src.hyperlink)


def copy_column_widths(src_ws: Worksheet, dst_ws: Worksheet, max_col: int) -> None:
    for c in range(1, max_col + 1):
        letter = openpyxl.utils.get_column_letter(c)
        width = src_ws.column_dimensions[letter].width
        if width:
            dst_ws.column_dimensions[letter].width = width


def to_time_value(v):
    if isinstance(v, datetime):
        return v.time().replace(second=0, microsecond=0)
    if isinstance(v, time):
        return v.replace(second=0, microsecond=0)
    if isinstance(v, (int, float)):
        minutes = round(float(v) * 24 * 60)
        return (datetime(1899, 12, 30) + timedelta(minutes=minutes)).time().replace(second=0, microsecond=0)
    if isinstance(v, str):
        s = v.strip()
        for fmt in ["%H:%M", "%H:%M:%S"]:
            try:
                return datetime.strptime(s, fmt).time()
            except ValueError:
                pass
    return v


def calc_alarm(start_value):
    t = to_time_value(start_value)
    if isinstance(t, time):
        dt = datetime.combine(date(2000, 1, 1), t) - timedelta(minutes=5)
        return dt.time().replace(second=0, microsecond=0)
    return None


def select_source_sheet(plan: str, d: date) -> str:
    return f"{plan}{WEEKEND_SUFFIX}" if is_weekend(d) else plan


def format_date_cell(cell, d: date) -> None:
    cell.value = datetime(d.year, d.month, d.day)
    cell.number_format = DATE_FORMAT


def build_workbook(uploaded_file, schedule_items: pd.DataFrame, clear_existing: bool, target_sheet_name: str) -> Tuple[BytesIO, List[Dict[str, str]]]:
    wb = openpyxl.load_workbook(uploaded_file)
    if target_sheet_name not in wb.sheetnames:
        raise ValueError(f"'{target_sheet_name}' 시트를 찾지 못했습니다. 첫 번째 출력 시트 이름을 확인하세요.")
    target_ws = wb[target_sheet_name]
    target_header_row, target_cols = find_header(target_ws)
    if clear_existing:
        clear_schedule_body(target_ws, target_header_row, target_cols)
        append_row = target_header_row + 1
    else:
        append_row = last_used_row(target_ws, target_cols) + 1
    logs: List[Dict[str, str]] = []
    source_cache: Dict[str, Tuple[int, Dict[str, int], List[int]]] = {}

    valid_rows = schedule_items.dropna(subset=["date", "plan"])
    for _, item in valid_rows.iterrows():
        d = item["date"]
        if isinstance(d, pd.Timestamp):
            d = d.date()
        elif isinstance(d, datetime):
            d = d.date()
        plan = str(item["plan"]).strip()
        src_sheet_name = select_source_sheet(plan, d)
        if src_sheet_name not in wb.sheetnames:
            raise ValueError(f"'{src_sheet_name}' 시트가 없습니다. 평일/주말용 6개 시트명을 확인하세요.")
        src_ws = wb[src_sheet_name]
        if src_sheet_name not in source_cache:
            header_row, src_cols = find_header(src_ws)
            rows = source_data_rows(src_ws, header_row, src_cols)
            if not rows:
                raise ValueError(f"'{src_sheet_name}' 시트에서 복사할 타임테이블 행을 찾지 못했습니다.")
            source_cache[src_sheet_name] = (header_row, src_cols, rows)
        _, src_cols, rows = source_cache[src_sheet_name]
        max_col = max(src_ws.max_column, target_ws.max_column, max(target_cols.values()), max(src_cols.values()))
        copy_column_widths(src_ws, target_ws, max_col)
        start_append_row = append_row
        for src_row in rows:
            copy_row_style_and_values(src_ws, target_ws, src_row, append_row, max_col)
            if "date" in target_cols:
                format_date_cell(target_ws.cell(append_row, target_cols["date"]), d)
            if "alarm" in target_cols and "start" in target_cols:
                alarm = calc_alarm(target_ws.cell(append_row, target_cols["start"]).value)
                if alarm is not None:
                    target_ws.cell(append_row, target_cols["alarm"]).value = alarm
                    target_ws.cell(append_row, target_cols["alarm"]).number_format = TIME_FORMAT
            if "start" in target_cols:
                target_ws.cell(append_row, target_cols["start"]).number_format = TIME_FORMAT
            append_row += 1
        logs.append({
            "날짜": d.strftime("%Y-%m-%d"),
            "선택 플랜": plan,
            "적용 시트": src_sheet_name,
            "입력 행수": str(append_row - start_append_row),
        })
    if target_ws.auto_filter:
        target_ws.auto_filter.ref = None
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output, logs


st.set_page_config(page_title="HMG 사이니지 스케줄 생성기", layout="wide")
st.title("HMG 사이니지 스케줄 생성기")
st.caption("엑셀을 직접 편집하지 않고 날짜와 플랜만 선택해 첫 번째 스케줄 시트에 타임테이블을 누적 입력합니다.")

with st.sidebar:
    st.header("1. 기준 엑셀")
    uploaded = st.file_uploader("6개 타임테이블 시트가 들어있는 엑셀 업로드", type=["xlsx"])
    target_sheet_name = st.text_input("출력 시트명", value="스케줄")
    clear_existing = st.checkbox("기존 스케줄 내용 삭제 후 새로 생성", value=True)
    st.divider()
    st.header("시트명 규칙")
    st.write("평일: Plan1, Plan2, Plan3")
    st.write("주말: Plan1 (주말), Plan2 (주말), Plan3 (주말)")

st.subheader("2. 날짜별 플랜 입력")
st.write("날짜 하나당 플랜 하나를 선택하세요. 토/일이면 주말 시트가 자동 적용됩니다.")

base_df = pd.DataFrame([
    {"date": date(2026, 4, 1), "plan": "Plan1"},
    {"date": date(2026, 4, 2), "plan": "Plan2"},
])
items = st.data_editor(
    base_df,
    num_rows="dynamic",
    use_container_width=True,
    column_config={
        "date": st.column_config.DateColumn("날짜", format="YYYY-MM-DD", required=True),
        "plan": st.column_config.SelectboxColumn("플랜", options=REQUIRED_PLANS, required=True),
    },
    hide_index=True,
)

preview = items.copy()
if not preview.empty and "date" in preview.columns:
    preview["적용 시트"] = preview.apply(
        lambda row: select_source_sheet(row["plan"], row["date"] if not isinstance(row["date"], pd.Timestamp) else row["date"].date())
        if pd.notna(row.get("date")) and pd.notna(row.get("plan")) else "",
        axis=1,
    )
    st.subheader("3. 적용 예정 미리보기")
    st.dataframe(preview, use_container_width=True, hide_index=True)

if st.button("엑셀 생성", type="primary", use_container_width=True):
    if uploaded is None:
        st.error("기준 엑셀 파일을 먼저 업로드하세요.")
    elif items.dropna(subset=["date", "plan"]).empty:
        st.error("날짜와 플랜을 최소 1개 이상 입력하세요.")
    else:
        try:
            output, logs = build_workbook(uploaded, items, clear_existing, target_sheet_name)
            st.success("생성 완료")
            st.dataframe(pd.DataFrame(logs), use_container_width=True, hide_index=True)
            st.download_button(
                "완성 엑셀 다운로드",
                data=output,
                file_name="signage_schedule_generated.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except Exception as e:
            st.error(str(e))
            st.info("시트명, 첫 번째 출력 시트명, 헤더명(날짜/알람시간/시작예정시각/프로그램명/렉처룸)을 확인하세요.")
